"""Exporta dados para Excel com formatação premium usando openpyxl."""
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import Any


class ExcelExporter:
    # Tema escuro premium
    HEADER_FILL = PatternFill(start_color="0D1117", end_color="0D1117", fill_type="solid")
    HEADER_FONT = Font(color="00D4FF", bold=True, size=11)  # cyan
    ROW_FILL_1 = PatternFill(start_color="161B22", end_color="161B22", fill_type="solid")
    ROW_FILL_2 = PatternFill(start_color="0D1117", end_color="0D1117", fill_type="solid")
    ROW_FONT = Font(color="E6EDF3", size=10)
    THIN_BORDER = Border(
        bottom=Side(style="thin", color="30363D"),
    )

    def export(self, data: list[dict], filename: str | None = None, title: str = "WebScrapy Export") -> str:
        wb = self._build_workbook(data, title)
        path = filename or f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(path)
        return path

    def export_bytes(self, data: list[dict], title: str = "WebScrapy Export") -> bytes:
        wb = self._build_workbook(data, title)
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def _build_workbook(self, data: list[dict], title: str) -> Workbook:
        wb = Workbook()
        ws = wb.active
        ws.title = title[:31]  # Excel max sheet name length

        if not data:
            ws["A1"] = "Sem dados para exportar"
            return wb

        headers = list(data[0].keys())

        # Header row
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header.upper().replace("_", " "))
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = self.THIN_BORDER

        # Data rows
        for row_idx, item in enumerate(data, 2):
            fill = self.ROW_FILL_1 if row_idx % 2 == 0 else self.ROW_FILL_2
            for col_idx, header in enumerate(headers, 1):
                value = item.get(header, "")
                if isinstance(value, (list, dict)):
                    value = str(value)
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.fill = fill
                cell.font = self.ROW_FONT
                cell.border = self.THIN_BORDER
                cell.alignment = Alignment(vertical="center", wrap_text=False)

        # Auto-fit columns
        for col_idx, header in enumerate(headers, 1):
            max_width = max(len(header), 10)
            for row_idx in range(2, min(len(data) + 2, 102)):
                cell_value = str(ws.cell(row=row_idx, column=col_idx).value or "")
                max_width = max(max_width, len(cell_value))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_width + 2, 50)

        # Freeze header row
        ws.freeze_panes = "A2"

        # Metadata sheet
        meta = wb.create_sheet("Metadata")
        meta["A1"] = "WebScrapy AI Platform Export"
        meta["A2"] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        meta["A3"] = f"Total de registros: {len(data)}"
        meta["A4"] = f"Colunas: {', '.join(headers)}"

        return wb
